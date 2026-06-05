import os
import pandas as pd
from typing import Optional
from langchain_core.tools import tool
from langchain_openai import ChatOpenAI
from langchain_core.messages import SystemMessage, HumanMessage
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from app.core.config import settings

@tool
def screen_abstracts_csv(csv_path: str, criteria: str, output_path: Optional[str] = None, feedback: Optional[str] = None) -> str:
    """
    Evaluates academic abstracts in a CSV file against specified inclusion criteria.
    
    The CSV must have 'title' and 'abstract' columns. 
    It generates an Excel file with color-coded rows (green for keep, red for reject)
    and an LLM-generated justification.
    
    Args:
        csv_path: Path to the input CSV file.
        criteria: Inclusion/exclusion criteria to screen against.
        output_path: Optional path for the output Excel file. Defaults to input_name.xlsx.
        feedback: Extra user considerations or specific instructions for screening.
    """
    if not os.path.exists(csv_path):
        return f"Error: File not found at {csv_path}"

    try:
        df = pd.read_csv(csv_path)
    except Exception as e:
        return f"Error reading CSV: {str(e)}"

    if 'title' not in df.columns or 'abstract' not in df.columns:
        return "Error: CSV must contain 'title' and 'abstract' columns."

    llm = ChatOpenAI(
        model=settings.OPENAI_MODEL,
        api_key=settings.OPENAI_API_KEY,
        temperature=0
    )

    results = []
    justifications = []
    decisions = []

    system_prompt = (
        "You are an expert academic screener. You will be given a title, an abstract, "
        "and inclusion criteria. Your task is to decide if the paper should be 'KEPT' or 'REJECTED' "
        "based on the criteria. Provide a concise justification for your decision."
    )
    if feedback:
        system_prompt += f"\n\nAdditional User Feedback/Considerations: {feedback}"
    
    system_prompt += (
        "\n\nFormat your response exactly as follows:"
        "\nDECISION: [KEEP/REJECT]"
        "\nJUSTIFICATION: [Your reason]"
    )

    for index, row in df.iterrows():
        user_content = (
            f"Criteria: {criteria}\n\n"
            f"Title: {row['title']}\n\n"
            f"Abstract: {row['abstract']}"
        )
        
        try:
            response = llm.invoke([
                SystemMessage(content=system_prompt),
                HumanMessage(content=user_content)
            ])
            content = response.content
            
            # Simple parsing
            decision = "REJECT"
            justification = "Error parsing LLM response."
            
            if "DECISION: KEEP" in content.upper():
                decision = "KEEP"
            elif "DECISION: REJECT" in content.upper():
                decision = "REJECT"
                
            if "JUSTIFICATION:" in content.upper():
                justification = content.upper().split("JUSTIFICATION:")[1].strip()
            
            decisions.append(decision)
            justifications.append(justification)
            
        except Exception as e:
            decisions.append("ERROR")
            justifications.append(f"LLM Error: {str(e)}")

    df['decision'] = decisions
    df['justification'] = justifications

    if not output_path:
        output_path = csv_path.rsplit('.', 1)[0] + "_screened.xlsx"

    # Save to Excel
    df.to_excel(output_path, index=False)

    # Apply formatting
    wb = load_workbook(output_path)
    ws = wb.active

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Assuming decision is in the last but one column (index len(df.columns) - 1 if zero-indexed)
    # df.to_excel puts headers in row 1. Data starts in row 2.
    decision_col_idx = df.columns.get_loc('decision') + 1 # 1-based for openpyxl

    for row_idx in range(2, ws.max_row + 1):
        decision_val = ws.cell(row=row_idx, column=decision_col_idx).value
        fill = None
        if decision_val == "KEEP":
            fill = green_fill
        elif decision_val == "REJECT":
            fill = red_fill
            
        if fill:
            for cell in ws[row_idx]:
                cell.fill = fill

    wb.save(output_path)

    return f"Screening complete. Results saved to {output_path}. Total papers processed: {len(df)}."
